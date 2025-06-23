import streamlit as st
import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook

df = pd.read_excel("Datenmodell.xlsx")

def produktkonfigurator():
    st.title("🛠️ Produktkonfigurator")

    daten_datei = "Datenmodell.xlsx"
    sheet_name_daten = "Daten"
    sheet_name_produkte = "Produkte"

    # Lade Daten-Sheet
    try:
        df_daten = pd.read_excel("Datenmodell.xlsx")
    except Exception as e:
        st.error(f"Fehler beim Laden der Datei: {e}")
        return

    # Lade oder initialisiere das Produkte-Sheet
    if os.path.exists(daten_datei):
        try:
            wb = load_workbook(daten_datei)
            if sheet_name_produkte in wb.sheetnames:
                df_produkte = pd.read_excel(daten_datei, sheet_name=sheet_name_produkte)
            else:
                df_produkte = pd.DataFrame(columns=["Produkt ID", "Produkt Name"])
        except Exception as e:
            st.error(f"Fehler beim Laden des Produkte-Blattes: {e}")
            return
    else:
        st.error("Excel-Datei nicht gefunden.")
        return

    # Produkt-Eingabe
    produkt_id = st.text_input("🆔 Produkt ID")
    produkt_name = st.text_input("📦 Produkt Name")

    if produkt_id and produkt_name:
        if produkt_id in df_produkte["Produkt ID"].astype(str).values:
            st.warning("❗Ein Produkt mit dieser ID existiert bereits!")
            return

        # Auswahl der gewünschten Kennzahlen
        alle_kennzahlen = df_daten["Name"].dropna().unique().tolist()
        ausgewaehlte_kennzahlen = st.multiselect(
            "Kennzahlen auswählen (suche möglich):",
            options=alle_kennzahlen,
            help="Wähle eine oder mehrere Kennzahlen aus",
        )

        # Werteingabe für die gewählten Kennzahlen
        benutzereingaben = {}
        for kennzahl in ausgewaehlte_kennzahlen:
            wert = st.text_input(f"Wert für '{kennzahl}':")
            benutzereingaben[kennzahl] = wert

        if st.button("💾 Produkt speichern"):
            # Neue Zeile mit Produktdaten
            neue_zeile = {
                "Produkt ID": produkt_id,
                "Produkt Name": produkt_name,
            }
            neue_zeile.update(benutzereingaben)

            # Aktuelle Tabelle aktualisieren
            df_produkte = pd.concat([df_produkte, pd.DataFrame([neue_zeile])], ignore_index=True)

            # Schreibe alle Daten in die Excel-Datei
            with pd.ExcelWriter(daten_datei, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df_produkte.to_excel(writer, sheet_name=sheet_name_produkte, index=False)
                df_daten.to_excel(writer, sheet_name=sheet_name_daten, index=False)  # Sicherstellen, dass "Daten" bleibt

            st.success("✅ Produkt erfolgreich gespeichert!")
    else:
        st.info("Bitte Produkt ID und Produkt Name eingeben.")
produktkonfigurator()
